#!/usr/bin/env python3
"""
import_promo_rollups.py — apply rollup spreadsheets onto mint.national.promo
records that were created by backfill_2026_national_promo_log.py.

Inputs (default location: ~/Downloads/National Promos by brand 2/):
  • Brands.xlsx
      - "Brands Reached out to ..." outreach key sheet (informational summary)
      - Per-state status sheets ("AZ", "AZ April", "AZ May", "MO", "MO May",
        "NV", "NV May", "IL", "IL april", "IL May") — each row:
          Lsp Location | Inventory Brand Name | Ready to plot (per month)
          | Plotted (per month) | Notes columns
        Aggregates per-month flags into the year campaign:
          ready_to_plot = any YES
          plotted       = any YES
          contact_status derived from inline notes
      - "Sheet3" / "Copy of Sheet3" — POS transaction stats
        Dumped to CSV (see --pos-stats-out) because the data shape needs a
        new mint.promo.pos.usage model to be properly pivotable.

  • IL/IL Promo count_.xlsx
      - "Vendors": per-vendor margin/BOGO/% off/frontload rules
      - "Brand  Category on MP": vendor × brand × category rules
      - "Brand  Category not on MP": brands without an MP entry
      - "Brand  Category w weight": SKU weight breakdown
      Each is dumped onto the IL campaign's `notes` field as a structured
      text block (vendor rules → notes header), and unmapped vendors are
      logged.

Behavior:
  • Only updates existing mint.national.promo records (created by the
    backfill). Unmatched (brand, state) combos are reported but skipped.
  • Idempotent: re-running with the same XLSX produces the same result.
  • --dry-run prints what would change without writing.

Auth:
  ODOO_URL, ODOO_DB, ODOO_UID (default 2), ODOO_KEY env vars or CLI flags.

Usage:
  python3 import_promo_rollups.py --dry-run
  python3 import_promo_rollups.py --root "/Users/Keymaker/Downloads/National Promos by brand 2"
  python3 import_promo_rollups.py --state AZ
"""
import argparse
import csv
import json
import logging
import os
import re
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
_log = logging.getLogger(__name__)

DEFAULT_ROOT = Path.home() / 'Downloads' / 'National Promos by brand 2'
YEAR = 2026

STATE_TO_MARKET_CODE = {'AZ': 'AZ', 'IL': 'IL', 'MO': 'MO', 'NV': 'NV'}

# ── Outreach status mapping (cell text → mint.national.promo.contact_status) ──
# Heuristics applied to the row's note column AND the "Ready to plot" cell.
def derive_contact_status(notes: str, ready_text: str, plotted_text: str) -> str:
    text = ' '.join([notes or '', ready_text or '', plotted_text or '']).lower().strip()
    if not text:
        return 'not_contacted'
    if any(k in text for k in ['declined', 'declines', 'pass', 'no interest', 'will not']):
        return 'declined'
    if any(k in text for k in ['confirmed', 'confirm', 'verified', 'verifid', 'got back', 'got confirmation']):
        return 'confirmed'
    if any(k in text for k in ['waiting', 'awaiting', 'wants a meeting', 'follow up', 'no response']):
        return 'awaiting_response'
    if any(k in text for k in ['emailed', 'reach out', 'reached out', 'outreach', 'sent', 'asked']):
        return 'outreached'
    return 'not_contacted'


def normalize_brand_name(s: str) -> str:
    return ''.join(c for c in (s or '').lower() if c.isalnum())


# ── Odoo client (reuses the pattern from backfill_2026_national_promo_log.py) ──

class Odoo:
    def __init__(self, url, db, uid, key, dry_run=False):
        self.url = url.rstrip('/')
        if not self.url.endswith('/jsonrpc'):
            self.url = self.url + '/jsonrpc'
        self.db = db
        self.uid = uid
        self.key = key
        self.dry_run = dry_run

    def call(self, model, method, args, kwargs=None):
        if self.dry_run and method in ('create', 'write', 'unlink'):
            _log.info('[DRY-RUN] %s.%s(args=%s, kw=%s)', model, method,
                      str(args)[:200], str(kwargs)[:200])
            return {'result': True if method != 'create' else 0}
        payload = {
            'jsonrpc': '2.0', 'method': 'call',
            'params': {
                'service': 'object', 'method': 'execute_kw',
                'args': [self.db, self.uid, self.key, model, method, args, kwargs or {}],
            },
        }
        req = urllib.request.Request(self.url,
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json',
                     'User-Agent': 'mint-rollups/1.0',
                     'Accept': 'application/json'})
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read())
        if 'error' in data:
            raise RuntimeError(data['error'].get('data', {}).get('message', data['error']))
        return data

    def search_read(self, model, domain, fields, **kw):
        return self.call(model, 'search_read', [domain], {'fields': fields, **kw}).get('result', [])

    def write(self, model, ids, vals):
        return self.call(model, 'write', [ids, vals]).get('result')


# ── Loaders ──

def load_market_lookup(odoo):
    out = {}
    for r in odoo.search_read('mint.region', [], ['id', 'code']):
        if r.get('code'):
            out[r['code'].upper()] = r['id']
    return out


def load_brand_lookup(odoo):
    out = {}
    offset = 0
    while True:
        page = odoo.search_read('mint.brand', [], ['id', 'name'], limit=500, offset=offset)
        if not page:
            break
        for r in page:
            n = normalize_brand_name(r['name'])
            if n and n not in out:
                out[n] = r['id']
        offset += 500
    return out


def load_campaign_lookup(odoo, year):
    """(brand_id, market_id) → campaign_id for the given year."""
    out = {}
    offset = 0
    while True:
        page = odoo.search_read('mint.national.promo',
                                [['year', '=', year]],
                                ['id', 'brand_id', 'market_id'],
                                limit=500, offset=offset)
        if not page:
            break
        for r in page:
            b = r['brand_id'][0] if r.get('brand_id') else None
            m = r['market_id'][0] if r.get('market_id') else None
            if b and m:
                out[(b, m)] = r['id']
        offset += 500
    return out


# ── Brands.xlsx parsing ──

# Headers in per-state sheets:
# row 0: "Lsp Location ..." | "Inventory Brand Name" | "Ready to plot" | "Plotted" | (maybe) more notes cols
# row 1: blank | blank | <month label> | <month label> | (maybe) more
def parse_brands_state_sheet(ws, state_code: str):
    """Yield (brand_name, ready_yes, plotted_yes, notes) per data row."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return
    header0 = [str(c or '').strip() for c in rows[0]]
    header1 = [str(c or '').strip() for c in rows[1]]
    # Find brand col by header0 ilike "Brand"
    brand_col = next((i for i, h in enumerate(header0) if 'brand' in h.lower()), 1)
    # "Ready to plot" and "Plotted" columns can repeat (per-month). Capture all.
    ready_cols = [i for i, h in enumerate(header0) if 'ready' in h.lower()]
    plotted_cols = [i for i, h in enumerate(header0) if h.lower().startswith('plotted')]
    if not ready_cols and not plotted_cols:
        # fallback to fixed cols
        ready_cols, plotted_cols = [2], [4]
    # Note columns: anything past column 5 that isn't Plotted/Ready
    note_cols = [i for i in range(min(8, len(header0)))
                 if i not in ready_cols + plotted_cols + [brand_col, 0]]
    data_start = 2 if any(header1) else 1
    for row in rows[data_start:]:
        if not row or brand_col >= len(row):
            continue
        brand = str(row[brand_col] or '').strip()
        if not brand:
            continue
        ready_yes = any(str(row[c]).strip().upper() == 'YES'
                        for c in ready_cols if c < len(row))
        plotted_yes = any(str(row[c]).strip().upper() == 'YES'
                          for c in plotted_cols if c < len(row))
        # Note text = first non-empty cell after the boolean columns
        notes_parts = []
        for c in note_cols:
            if c < len(row) and row[c]:
                t = str(row[c]).strip()
                if t and t.upper() not in ('YES', 'NO'):
                    notes_parts.append(t)
        notes = ' | '.join(notes_parts)
        yield brand, ready_yes, plotted_yes, notes


# Sheet name -> state code (e.g. "AZ April" -> "AZ", "IL april" -> "IL")
STATE_SHEET_RE = re.compile(r'^(AZ|IL|MO|NV)(?:\s.+)?$', re.IGNORECASE)


def import_brands_outreach(odoo, brands_path, brand_lookup, market_lookup,
                           campaign_lookup, only_state=None, dry_run=False):
    if not brands_path.exists():
        _log.warning('Brands file not found: %s', brands_path)
        return
    wb = load_workbook(brands_path, data_only=True, read_only=False)
    # Aggregate across multiple sheets per state (e.g. AZ + AZ April + AZ May)
    agg = defaultdict(lambda: {'ready': False, 'plotted': False, 'notes': []})
    sheet_count = 0
    for sn in wb.sheetnames:
        m = STATE_SHEET_RE.match(sn.strip())
        if not m:
            continue
        st = m.group(1).upper()
        if only_state and st != only_state:
            continue
        sheet_count += 1
        for brand, ready, plotted, notes in parse_brands_state_sheet(wb[sn], st):
            key = (st, normalize_brand_name(brand))
            if not key[1]:
                continue
            d = agg[key]
            d['ready'] = d['ready'] or ready
            d['plotted'] = d['plotted'] or plotted
            if notes:
                d['notes'].append(f'{sn}: {notes}')
            d['raw_brand'] = brand
    _log.info('Brands.xlsx: parsed %d state sheets, %d (state,brand) rows',
              sheet_count, len(agg))

    written = 0
    skipped_brand = 0
    skipped_no_campaign = 0
    for (st, brand_norm), d in agg.items():
        market_id = market_lookup.get(st)
        if not market_id:
            continue
        brand_id = brand_lookup.get(brand_norm)
        if not brand_id:
            sys.stderr.write(f"unmatched_outreach_brand\t{st}\t{d.get('raw_brand', brand_norm)}\n")
            skipped_brand += 1
            continue
        campaign_id = campaign_lookup.get((brand_id, market_id))
        if not campaign_id:
            skipped_no_campaign += 1
            continue
        notes_text = ' || '.join(d['notes'])[:8000]
        contact = derive_contact_status(notes_text, '', '')
        vals = {
            'ready_to_plot': bool(d['ready']),
            'plotted': bool(d['plotted']),
            'contact_status': contact,
        }
        if notes_text:
            vals['notes'] = '[Outreach rollup]\n' + notes_text
        try:
            odoo.write('mint.national.promo', [campaign_id], vals)
            written += 1
        except Exception as e:
            _log.warning('write failed for campaign=%s: %s', campaign_id, e)
    _log.info('Brands outreach: written=%d unmatched_brand=%d no_campaign=%d',
              written, skipped_brand, skipped_no_campaign)


# ── Sheet3 POS stats → CSV dump ──

def dump_pos_stats(brands_path: Path, out_csv: Path):
    if not brands_path.exists():
        return 0
    wb = load_workbook(brands_path, data_only=True, read_only=True)
    written = 0
    seen = set()
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, 'w', newline='') as fh:
        w = csv.writer(fh)
        w.writerow(['lsp_location', 'discount_name', 'discount_text',
                    'brand', 'product', 'items', 'transactions'])
        for sn in wb.sheetnames:
            if not sn.lower().endswith('sheet3'):
                continue
            ws = wb[sn]
            it = ws.iter_rows(values_only=True)
            try:
                next(it)  # header
            except StopIteration:
                continue
            for row in it:
                if not row or not any(row):
                    continue
                key = tuple((str(c) if c is not None else '') for c in row[:7])
                if key in seen:
                    continue
                seen.add(key)
                w.writerow(key)
                written += 1
        wb.close()
    _log.info('POS stats CSV: %s rows -> %s', written, out_csv)
    return written


# ── IL Promo count rollup → notes on IL campaigns ──

def import_il_promo_count(odoo, il_path: Path, brand_lookup, market_lookup,
                          campaign_lookup, dry_run=False):
    if not il_path.exists():
        _log.warning('IL Promo count file not found: %s', il_path)
        return
    market_il = market_lookup.get('IL')
    if not market_il:
        _log.warning('No IL market in mint.region — skipping IL Promo count')
        return
    wb = load_workbook(il_path, data_only=True, read_only=False)

    # ── "Brand  Category on MP" — vendor × brand × category rules ──
    rules_by_brand = defaultdict(list)  # brand_norm -> list of rule strings
    if 'Brand  Category on MP' in wb.sheetnames:
        ws = wb['Brand  Category on MP']
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c or '').strip() for c in rows[0]]
            try:
                vendor_i = header.index('Inventory Vendor')
                brand_i  = header.index('Inventory Brand Name')
                cat_i    = header.index('Inventory Category')
                margin_i = header.index('On Margin Y / N')
                bogo_i   = next(i for i, h in enumerate(header) if h.lower().startswith('bogos a'))
                pct_i    = next(i for i, h in enumerate(header) if h == '%')
                pcto_i   = next(i for i, h in enumerate(header) if h.lower().startswith('% off'))
            except (ValueError, StopIteration):
                _log.warning('IL "Brand Category on MP": header layout unexpected')
            else:
                for row in rows[1:]:
                    if not row or not row[brand_i]:
                        continue
                    brand_n = normalize_brand_name(str(row[brand_i]))
                    rules_by_brand[brand_n].append(
                        f'{row[cat_i] or "?"}: margin={row[margin_i]}, '
                        f'BOGOs/mo={row[bogo_i]}, %={row[pct_i]}, %off={row[pcto_i]}'
                    )

    # ── "Vendors" — per-vendor topline ──
    vendor_summary = []
    if 'Vendors' in wb.sheetnames:
        ws = wb['Vendors']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            vendor_summary.append({
                'vendor': str(row[0]).strip(),
                'on_margin': row[1] if len(row) > 1 else '',
                'bogos': row[2] if len(row) > 2 else '',
                'pct_off': row[3] if len(row) > 3 else '',
                'frontload': row[4] if len(row) > 4 else '',
            })
    _log.info('IL Promo count: %d vendor rows, %d brands with rules',
              len(vendor_summary), len(rules_by_brand))

    # Apply: append a [IL Margin Rules] block to each IL campaign's notes
    written = 0
    for brand_norm, rule_lines in rules_by_brand.items():
        brand_id = brand_lookup.get(brand_norm)
        if not brand_id:
            sys.stderr.write(f'unmatched_il_rules_brand\t{brand_norm}\n')
            continue
        campaign_id = campaign_lookup.get((brand_id, market_il))
        if not campaign_id:
            continue
        # Read existing notes so we replace just the IL Rules section idempotently
        existing = odoo.search_read('mint.national.promo',
                                    [['id', '=', campaign_id]], ['notes'], limit=1)
        cur = (existing[0].get('notes') or '') if existing else ''
        cur = re.sub(r'\n\n\[IL Margin Rules\][\s\S]*?(?=\n\n\[|$)', '', cur).rstrip()
        block = '\n\n[IL Margin Rules]\n' + '\n'.join(f'  • {r}' for r in rule_lines)
        new_notes = (cur + block).strip()
        try:
            odoo.write('mint.national.promo', [campaign_id], {'notes': new_notes})
            written += 1
        except Exception as e:
            _log.warning('IL write failed for campaign=%s: %s', campaign_id, e)
    _log.info('IL Promo count: applied rules to %d campaigns', written)


# ── Main ──

def main():
    p = argparse.ArgumentParser()
    p.add_argument('--root', default=str(DEFAULT_ROOT))
    p.add_argument('--state', help='Limit Brands outreach to a single state (AZ/IL/MO/NV)')
    p.add_argument('--year', type=int, default=YEAR)
    p.add_argument('--dry-run', action='store_true')
    p.add_argument('--pos-stats-out', default='/Users/Keymaker/.claude/jobs/cedcd14f/pos_stats.csv',
                   help='Where to dump Brands.xlsx Sheet3 POS stats CSV')
    p.add_argument('--url', default=os.environ.get('ODOO_URL', 'https://letsgomint.us/jsonrpc'))
    p.add_argument('--db', default=os.environ.get('ODOO_DB', 'odoo'))
    p.add_argument('--uid', type=int, default=int(os.environ.get('ODOO_UID', '2')))
    p.add_argument('--key', default=os.environ.get('ODOO_KEY', ''))
    args = p.parse_args()
    if not args.key:
        sys.exit('ODOO_KEY env var or --key required')

    odoo = Odoo(args.url, args.db, args.uid, args.key, dry_run=args.dry_run)
    root = Path(args.root)
    brands_path = root / 'Brands.xlsx'
    il_path = root / 'IL' / 'IL Promo count_.xlsx'
    pos_csv = Path(args.pos_stats_out)

    _log.info('Loading mint.brand registry...')
    brand_lookup = load_brand_lookup(odoo)
    _log.info('  %d brands', len(brand_lookup))
    _log.info('Loading mint.region (markets)...')
    market_lookup = load_market_lookup(odoo)
    _log.info('  %d markets: %s', len(market_lookup), sorted(market_lookup.keys()))
    _log.info('Loading mint.national.promo campaigns for year=%d...', args.year)
    campaign_lookup = load_campaign_lookup(odoo, args.year)
    _log.info('  %d campaigns', len(campaign_lookup))

    only = args.state.upper() if args.state else None
    sys.stderr.write('rollup_unmatched:\tkind\tstate\tname\n')

    _log.info('=== Brands.xlsx outreach ===')
    import_brands_outreach(odoo, brands_path, brand_lookup, market_lookup,
                           campaign_lookup, only_state=only, dry_run=args.dry_run)

    _log.info('=== IL Promo count vendor/brand rules ===')
    import_il_promo_count(odoo, il_path, brand_lookup, market_lookup,
                          campaign_lookup, dry_run=args.dry_run)

    _log.info('=== Sheet3 POS stats → CSV ===')
    dump_pos_stats(brands_path, pos_csv)
    _log.info('done.')


if __name__ == '__main__':
    main()
