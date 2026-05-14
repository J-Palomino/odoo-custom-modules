#!/usr/bin/env python3
"""
backfill_2026_national_promo_log.py — import per-brand 2026 promo XLSXes
into mint.national.promo + mint.brand.calendar.entry.

Reads:
    ~/Downloads/National Promos by brand/{AZ,IL,MO,NV}/2026/*.xlsx

Behavior:
  - One mint.national.promo per (brand, market, year=2026), idempotent.
  - One mint.brand.calendar.entry per (brand, date, market, sku_or_category, campaign).
  - Brand resolved by case-insensitive normalized name vs. mint.brand. Unmatched
    brands are logged to a TSV on stderr and the whole file is skipped (the vendor
    manager curates mint.brand).
  - Cell content (e.g. "40% off", "2 for $59") is stored verbatim in `promo_text`
    AND parsed into (discount_type, discount_value, original_price) where possible.
  - Idempotent on the calendar entry's unique constraint.

Auth:
    JSON-RPC to Odoo. URL/DB/UID/API key read from env:
      ODOO_URL, ODOO_DB, ODOO_UID, ODOO_KEY
    Or pass --url, --db, --uid, --key.

Usage:
    python3 backfill_2026_national_promo_log.py --dry-run
    python3 backfill_2026_national_promo_log.py --state AZ
    python3 backfill_2026_national_promo_log.py
"""
import argparse
import json
import logging
import os
import re
import sys
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
_log = logging.getLogger(__name__)

DEFAULT_ROOT = Path.home() / 'Downloads' / 'National Promos by brand'

STATE_TO_MARKET_CODE = {
    'AZ': 'AZ',
    'IL': 'IL',
    'MO': 'MO',
    'NV': 'NV',
}

# Sheet name patterns. Original AZ/MO/IL files: "Jan promos", "Feb 2026", "May 2026".
# NV files use per-store sheets: "Paradise May", "Rainbow Jan", "Promos Jan - Rainbow",
# "April Both Stores", "Feb 2026 Both Locations", etc. Match any sheet that contains a
# month name as a standalone word; the data layout (Date column at index 3) is the same.
MONTH_NAMES = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
               'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
SHEET_RE = re.compile(
    r'\b(jan(?:uary)?|feb(?:ruary)?|march|mar|apr(?:il)?|may|june|jun|'
    r'july|jul|aug(?:ust)?|sept?(?:ember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b',
    re.IGNORECASE,
)


def normalize_brand_name(s: str) -> str:
    return ''.join(c for c in (s or '').lower() if c.isalnum())


def parse_brand_from_filename(name: str) -> str:
    base = name.rsplit('.', 1)[0]
    base = re.sub(r'\s*2026[_]?\s*', ' ', base)
    base = re.sub(r'\s*(AZ|IL|MO|NV)\s*$', ' ', base, flags=re.IGNORECASE)
    base = base.replace('_', ' ').strip()
    return base


PRICE_RE = re.compile(r'\$?(\d+(?:\.\d+)?)')


def parse_promo_text(text: str):
    """Return (discount_type, discount_value, original_price) or (None, None, None)."""
    if not text:
        return (None, None, None)
    t = re.sub(r'\s+', ' ', str(text).replace('\n', ' | ')).strip()
    tl = t.lower()

    # was/now with "each": "$40 $24 Each | 40% Off"
    m = re.search(r'\$(\d+(?:\.\d+)?)\s+\$(\d+(?:\.\d+)?)\s+each', t, re.IGNORECASE)
    if m:
        return ('price', float(m.group(2)), float(m.group(1)))

    # was/now compact: "$40 - $24"
    m = re.search(r'^\$(\d+(?:\.\d+)?)\s*[-–]\s*\$(\d+(?:\.\d+)?)', t)
    if m:
        return ('price', float(m.group(2)), float(m.group(1)))

    # BOGO / BxGy
    if 'bogo' in tl or re.search(r'b\d+g\d+', tl):
        return ('bogo', None, None)

    # multi-buy: "2 for $59"
    multi = re.findall(r'(\d+)\s*for\s*\$(\d+(?:\.\d+)?)', t, re.IGNORECASE)
    if multi:
        best = max(multi, key=lambda x: float(x[1]))
        return ('bundle', float(best[1]), None)

    # percent off
    m = re.search(r'(\d+(?:\.\d+)?)\s*%\s*off', tl)
    if m:
        return ('percent', float(m.group(1)), None)

    # only $X
    m = re.search(r'only\s*\$(\d+(?:\.\d+)?)', tl)
    if m:
        return ('price', float(m.group(1)), None)

    # fixed each: "$5 Each"
    m = re.match(r'^\$(\d+(?:\.\d+)?)\s+each', t, re.IGNORECASE)
    if m:
        return ('price', float(m.group(1)), None)

    return (None, None, None)


# ── Odoo JSON-RPC client ──

class Odoo:
    def __init__(self, url, db, uid, key, dry_run=False):
        self.url = url.rstrip('/')
        if not self.url.endswith('/jsonrpc'):
            self.url = self.url + '/jsonrpc'
        self.db = db
        self.uid = uid
        self.key = key
        self.dry_run = dry_run
        self.headers = {
            'Content-Type': 'application/json',
            'User-Agent': 'mint-backfill/1.0',
        }

    def call(self, model, method, args, kwargs=None):
        if self.dry_run and method in ('create', 'write', 'unlink'):
            _log.info('[DRY-RUN] %s.%s(args=%s, kw=%s)', model, method, args, kwargs)
            return {'result': True if method != 'create' else 0}
        payload = {
            'jsonrpc': '2.0',
            'method': 'call',
            'params': {
                'service': 'object',
                'method': 'execute_kw',
                'args': [self.db, self.uid, self.key, model, method, args, kwargs or {}],
            },
        }
        req = urllib.request.Request(
            self.url,
            data=json.dumps(payload).encode('utf-8'),
            headers=self.headers,
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read())
        if 'error' in data:
            raise RuntimeError(data['error'].get('data', {}).get('message', data['error']))
        return data

    def search_read(self, model, domain, fields, **kw):
        return self.call(model, 'search_read', [domain], {'fields': fields, **kw}).get('result', [])

    def create(self, model, vals):
        return self.call(model, 'create', [vals]).get('result')

    def write(self, model, ids, vals):
        return self.call(model, 'write', [ids, vals]).get('result')


# ── Main backfill ──

def load_brand_lookup(odoo):
    """Build normalized brand name → mint.brand.id map (paginated)."""
    out = {}
    offset = 0
    while True:
        page = odoo.search_read('mint.brand', [], ['id', 'name'], limit=500, offset=offset)
        if not page:
            break
        for rec in page:
            norm = normalize_brand_name(rec['name'])
            if norm and norm not in out:
                out[norm] = rec['id']
        offset += 500
    return out


def load_market_lookup(odoo):
    """Build state code → mint.region.id map."""
    regs = odoo.search_read('mint.region', [], ['id', 'code', 'name'])
    out = {}
    for r in regs:
        if r.get('code'):
            out[r['code'].upper()] = r['id']
    return out


def get_or_create_campaign(odoo, brand_id, market_id, year, source_file):
    existing = odoo.search_read(
        'mint.national.promo',
        [['brand_id', '=', brand_id], ['market_id', '=', market_id], ['year', '=', year]],
        ['id'],
        limit=1,
    )
    if existing:
        return existing[0]['id']
    return odoo.create('mint.national.promo', {
        'brand_id': brand_id,
        'market_id': market_id,
        'year': year,
        'source_file': source_file,
    })


def get_existing_entry_keys(odoo, campaign_id):
    """Return a set of (date_iso, sku_or_category) tuples already present."""
    rows = odoo.search_read(
        'mint.brand.calendar.entry',
        [['campaign_id', '=', campaign_id]],
        ['date', 'sku_or_category'],
        limit=10000,
    )
    return {(r['date'], (r.get('sku_or_category') or '')) for r in rows}


def process_workbook(odoo, path, brand_lookup, market_id, state_code, dry_run):
    brand_text = parse_brand_from_filename(path.name)
    brand_norm = normalize_brand_name(brand_text)
    brand_id = brand_lookup.get(brand_norm)
    if not brand_id:
        # try first word
        first = brand_text.split()[0] if brand_text else ''
        brand_id = brand_lookup.get(normalize_brand_name(first))
    if not brand_id:
        sys.stderr.write(f"unmatched_brand\t{state_code}\t{path.name}\t{brand_text}\n")
        return 0, 0

    rel_path = f"{state_code}/2026/{path.name}"
    campaign_id = get_or_create_campaign(
        odoo, brand_id, market_id, 2026, rel_path,
    )

    existing = get_existing_entry_keys(odoo, campaign_id) if not dry_run else set()

    wb = load_workbook(path, data_only=True, read_only=False)
    created = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        if not SHEET_RE.search(sheet_name.strip()):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        header = rows[1]
        if not header:
            continue
        # cols 0-3 = Weeks/Holidays/Day/Date (order varies — find by content)
        date_col = None
        for idx, val in enumerate(header[:6]):
            if isinstance(val, str) and val.strip().lower() == 'date':
                date_col = idx
                break
        if date_col is None:
            date_col = 3  # convention fallback
        sku_columns = [(idx, str(val).strip())
                       for idx, val in enumerate(header)
                       if idx > date_col and val and str(val).strip()]
        if not sku_columns:
            continue

        for row in rows[2:]:
            if not row or date_col >= len(row):
                continue
            row_date = row[date_col]
            if not row_date:
                continue
            # Coerce to date
            if hasattr(row_date, 'date'):
                row_date = row_date.date()
            iso = row_date.isoformat() if hasattr(row_date, 'isoformat') else str(row_date)

            for col_idx, sku in sku_columns:
                if col_idx >= len(row):
                    continue
                cell = row[col_idx]
                if not cell or not str(cell).strip():
                    continue
                key = (iso, sku)
                if key in existing:
                    skipped += 1
                    continue
                promo_text = str(cell).strip()
                dt, dv, op = parse_promo_text(promo_text)
                vals = {
                    'brand_id': brand_id,
                    'market_id': market_id,
                    'date': iso,
                    'campaign_id': campaign_id,
                    'sku_or_category': sku,
                    'promo_text': promo_text,
                    'state': 'tentative',
                    'is_published': False,
                }
                if dt is not None:
                    vals['discount_type'] = dt
                if dv is not None:
                    vals['discount_value'] = dv
                if op is not None:
                    vals['original_price'] = op
                try:
                    odoo.create('mint.brand.calendar.entry', vals)
                    created += 1
                    existing.add(key)
                except Exception as e:
                    _log.warning('create failed for %s %s %s: %s', brand_text, iso, sku, e)

    return created, skipped


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--root', default=str(DEFAULT_ROOT),
                   help=f'Promo XLSX root directory (default: {DEFAULT_ROOT})')
    p.add_argument('--state', help='Limit to a single state (AZ, IL, MO, NV)')
    p.add_argument('--dry-run', action='store_true', help='Log writes without committing')
    p.add_argument('--url', default=os.environ.get('ODOO_URL', 'https://letsgomint.us/jsonrpc'))
    p.add_argument('--db', default=os.environ.get('ODOO_DB', 'odoo'))
    p.add_argument('--uid', type=int, default=int(os.environ.get('ODOO_UID', '2')))
    p.add_argument('--key', default=os.environ.get('ODOO_KEY', ''))
    args = p.parse_args()

    if not args.key:
        sys.exit('ODOO_KEY env var or --key required')

    odoo = Odoo(args.url, args.db, args.uid, args.key, dry_run=args.dry_run)
    _log.info('Loading mint.brand registry...')
    brand_lookup = load_brand_lookup(odoo)
    _log.info('  %d brands', len(brand_lookup))
    _log.info('Loading mint.region (markets)...')
    market_lookup = load_market_lookup(odoo)
    _log.info('  %d markets', len(market_lookup))

    root = Path(args.root)
    states = [args.state.upper()] if args.state else list(STATE_TO_MARKET_CODE.keys())
    total_created = 0
    total_skipped = 0
    sys.stderr.write('unmatched_brand_log:\tstate\tfilename\tbrand_text\n')

    for state in states:
        market_code = STATE_TO_MARKET_CODE.get(state)
        if not market_code or market_code not in market_lookup:
            _log.warning('state %s: no matching mint.region — skipping', state)
            continue
        market_id = market_lookup[market_code]
        state_dir = root / state / '2026'
        if not state_dir.exists():
            _log.warning('state %s: no dir at %s — skipping', state, state_dir)
            continue
        xlsxes = sorted(p for p in state_dir.glob('*.xlsx') if not p.name.startswith('~$'))
        _log.info('state %s: %d XLSX files', state, len(xlsxes))
        for path in xlsxes:
            created, skipped = process_workbook(
                odoo, path, brand_lookup, market_id, state, args.dry_run,
            )
            total_created += created
            total_skipped += skipped
            _log.info('  %-50s created=%d skipped=%d', path.name, created, skipped)

    _log.info('done. created=%d skipped=%d', total_created, total_skipped)


if __name__ == '__main__':
    main()
