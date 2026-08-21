#!/usr/bin/env python3
"""
import_ptl_spreadsheet.py — load the monthly PTL "Store Copy" workbooks into
`mint.ptl.day` / `mint.ptl.deal` so the PTL Calendar reflects what the markets
are actually running, and so deals carry the keys needed to reconcile against
Dutchie.

Why this exists
---------------
The PTL is maintained in Google Sheets (one workbook per market per month, one
tab per calendar day). Odoo's copy has been drifting badly: for August 2026 the
spreadsheets hold 14,338 deal rows across AZ/IL/MO/NV while Odoo's August PTL
has 3,380 deals, and Arizona alone went from a 6,212-deal bulk load in May 2026
to *zero* deals created in August.

Worse, the fields needed to match a PTL deal to a Dutchie discount are the ones
being dropped in transfer. Measured on Odoo's August PTL vs the spreadsheets:

                       spreadsheet     Odoo PTL
    brand                   100%          57.1%
    product name            100%           0.1%
    weight                 96.6%          34.4%

So this importer's job is not just "create rows" — it is to carry `brand_id`
and `weight_value`/`weight_unit` across, because brand is the only exact join
key that both sides share (`mint.ptl.deal.brand_id` and `mint.discount.brand_ids`
both point at `mint.brand`), and weight is the key that becomes usable once
`mint.discount.weight_ids` is backfilled from `weight_restrictions_raw` — today
that field is 0% populated across all 1,477 Dutchie-sourced discounts.

The workbook's "Products" column is a form descriptor, not a SKU, so it is
stored verbatim on `description` rather than force-fitted onto
`explicit_product_ids` — see NOTE ON PRODUCTS below.

NO FUZZY MATCHING. Brands resolve by exact name or exact alias only. Anything
that does not resolve is left empty and reported, never guessed. This mirrors
the deal-pipeline rule that removed fuzzy matching in 2026-07: the sanctioned
mapping lives in data (`mint.brand.aliases`), and new spellings get added there
rather than inferred here.

Workbook shape (verified against all four August workbooks)
-----------------------------------------------------------
  row 1  market banner   e.g. "Mint — Arizona · Product ..."
  row 2  long date       e.g. "Friday, August 14, 2026"
  row 4  header          Category | Brand | Products | Format | Frequency | Deal | Locations
  row 5+ data            section header rows carry a Category but no Brand and are skipped

A deal that runs on several days appears on each of those day tabs. This script
upserts ONE `mint.ptl.deal` per distinct deal and links it to every day tab it
appears on, which is exactly what the `day_ids` many2many models. The
"Frequency" column is therefore descriptive only and is not parsed.

Idempotency
-----------
Re-running must not duplicate. Two independent guards:

  * Days   — searched by (date, market_id) and the LOWEST existing id wins.
             `mint.ptl.day` declares `unique(date, market_id)` in Python but
             that constraint is NOT installed in the database (verified: the
             table has only its 5 foreign keys). 96 duplicate (market, date)
             rows already exist, so the constraint cannot be relied on and this
             script must never create a day that already exists.
  * Deals  — keyed by a deterministic external id (`ir.model.data`) built from
             market + brand + product + format + deal text. Same row on a later
             run resolves to the same record and is updated, not re-created.

Safety
------
Deals are created `state='pending'`. Publishing to Dutchie is an explicit
`action_publish()` on `mint.ptl.day`; nothing here triggers a POS write.

Auth
----
ODOO_URL, ODOO_DB, ODOO_UID (default 2), ODOO_KEY env vars or CLI flags.

Usage
-----
  # inspect only (default — writes nothing)
  python3 import_ptl_spreadsheet.py --file "~/Downloads/Store Copy of AZ_PTL_Aug2026.xlsx"

  # actually write
  python3 import_ptl_spreadsheet.py --file "..." --commit

  # a whole month across markets
  python3 import_ptl_spreadsheet.py --file a.xlsx --file b.xlsx --commit
"""
import argparse
import hashlib
import json
import logging
import os
import re
import sys
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

_log = logging.getLogger('ptl_import')

IMPORT_MODULE = '__ptl_spreadsheet_import__'

# Filename prefix -> mint.region name. Explicit; never inferred from content.
MARKET_BY_CODE = {
    'AZ': 'Arizona', 'FL': 'Florida', 'IL': 'Illinois',
    'MI': 'Michigan', 'MO': 'Missouri', 'NV': 'Nevada',
}

# Spreadsheet "Format" -> mint.ptl.deal.discount_type. A declared mapping, not a
# guess: anything absent here is left unset and reported so a human can decide.
# "Mix & Match" and "Custom" are deliberately absent — they describe how the
# deal is rung up, not a discount arithmetic, and mapping them to `bundle`
# would silently misstate the deal.
FORMAT_TO_DISCOUNT_TYPE = {
    '% off': 'percent',
    'fixed price': 'price',
    'bogo': 'bogo',
    'bundle': 'bundle',
    'clearance': 'clearance',
}

VALID_WEIGHT_UNITS = {'g', 'mg', 'oz', 'ct'}
WEIGHT_RE = re.compile(r'(\d+(?:\.\d+)?)\s*(g|mg|oz|ct)\b', re.I)
MONTHS = ('january february march april may june july august september '
          'october november december').split()
DATE_RE = re.compile(r'(' + '|'.join(MONTHS) + r')\s+(\d{1,2}),?\s+(\d{4})', re.I)


class Odoo:
    """JSON-RPC client — same shape as import_promo_rollups.py."""

    def __init__(self, url, db, uid, key, dry_run=True):
        self.url = url.rstrip('/')
        if not self.url.endswith('/jsonrpc'):
            self.url += '/jsonrpc'
        self.db, self.uid, self.key = db, uid, key
        self.dry_run = dry_run
        self._fake_id = 0

    def call(self, model, method, args, kwargs=None):
        if self.dry_run and method in ('create', 'write', 'unlink'):
            _log.debug('[DRY-RUN] %s.%s %s', model, method, str(args)[:160])
            if method == 'create':
                self._fake_id -= 1          # negative ids never collide with real ones
                return self._fake_id
            return True
        payload = {
            'jsonrpc': '2.0', 'method': 'call',
            'params': {
                'service': 'object', 'method': 'execute_kw',
                'args': [self.db, self.uid, self.key, model, method, args, kwargs or {}],
            },
        }
        req = urllib.request.Request(
            self.url, data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json',
                     'User-Agent': 'mint-ptl-import/1.0',
                     'Accept': 'application/json'})
        with urllib.request.urlopen(req, timeout=180) as resp:
            data = json.loads(resp.read())
        if 'error' in data:
            raise RuntimeError(data['error'].get('data', {}).get('message', data['error']))
        return data.get('result')

    def search_read(self, model, domain, fields, **kw):
        return self.call(model, 'search_read', [domain], {'fields': fields, **kw}) or []

    def search(self, model, domain, **kw):
        return self.call(model, 'search', [domain], kw) or []

    def create(self, model, vals):
        return self.call(model, 'create', [vals])

    def write(self, model, ids, vals):
        return self.call(model, 'write', [ids, vals])


# ── exact-only resolvers ─────────────────────────────────────────────────────

# Group/collapse glyphs some workbooks prefix onto the Brand cell (the July
# AZ workbook writes "▼ Sofa King"). These are spreadsheet UI markers, not part
# of the brand name, so they are stripped before an exact lookup. Only these
# specific glyphs are removed — general leading punctuation is NOT stripped,
# because "&Shine" and "(the) Essence" are real brand names.
MARKER_RE = re.compile(r'^[▶▼▸▾►◀‣•]+\s*')


def norm_key(s):
    """Casefold + collapse whitespace + drop a leading collapse glyph. This is
    normalisation for an EXACT lookup (so "Alien Labs" == "alien  labs" and
    "▼ Sofa King" == "Sofa King"), not fuzzy matching — no token dropping, no
    substring or similarity logic."""
    s = MARKER_RE.sub('', (s or '').strip())
    return re.sub(r'\s+', ' ', s.strip()).casefold()


def build_brand_index(odoo):
    """name -> id and each alias -> id. Exact keys only.

    `mint.brand.aliases` is a free-text field; observed values are one alias per
    line, though comma-separated appears too, so both separators are split.

    `mint.brand` contains ~45 case-variant duplicate name groups (`STIIIZY` id 2
    vs `Stiiizy` id 931, `drip` id 13 vs `Drip` id 925, ...). In every observed
    pair exactly ONE side carries a `dutchie_brand_id` and the other is empty,
    so the Dutchie-linked record wins: this script exists to reconcile PTL deals
    against Dutchie discounts, and the brand carrying Dutchie's own identifier
    is by definition the right target. That is a deterministic tie-break on a
    real signal, not a similarity guess. If a collision is still ambiguous
    (both sides linked, or neither), the key is dropped and reported so a human
    resolves it in the data."""
    by_key = defaultdict(list)
    for b in odoo.search_read('mint.brand', [], ['name', 'aliases', 'dutchie_brand_id']):
        keys = [b['name']]
        if b.get('aliases'):
            keys += re.split(r'[\n,]+', b['aliases'])
        for raw in keys:
            k = norm_key(raw)
            if k:
                by_key[k].append(b)

    idx, collisions = {}, set()
    for k, cands in by_key.items():
        ids = {c['id'] for c in cands}
        if len(ids) == 1:
            idx[k] = cands[0]['id']
            continue
        linked = [c for c in cands if c.get('dutchie_brand_id')]
        if len({c['id'] for c in linked}) == 1:
            idx[k] = linked[0]['id']
        else:
            collisions.add(k)
    return idx, collisions


# NOTE ON PRODUCTS — why `explicit_product_ids` is not populated here.
#
# The workbook's "Products" column is a product-FORM + weight descriptor, not a
# SKU or a `product.template` name. Sampled values: "Flower 3.5g",
# "Gummies 100mg", "Cartridge Distillate 1g", "Premium Flower 7.77g". Exact
# matching those against `product.template.name` resolves 0% — correctly, since
# no product is named "Flower 3.5g".
#
# Rather than fuzzy-match a descriptor onto SKUs (which is exactly the guessing
# this project removed from the deal pipeline in 2026-07, and which the
# `deal_mixins.py` brand/category fallback already does damage with), the
# descriptor is stored verbatim on `description` — a field that is currently
# 0% populated on the August PTL. That keeps 100% of the spreadsheet's product
# information queryable, and leaves `explicit_product_ids` empty and honest.
#
# The usable exact join key is therefore brand + weight + category, with the
# descriptor available for a human or a later SKU-resolution pass.


def parse_weight(text):
    """('3.5','g') from 'Flower 3.5g'. Returns (None, None) when absent."""
    m = WEIGHT_RE.search(text or '')
    if not m:
        return None, None
    unit = m.group(2).lower()
    if unit not in VALID_WEIGHT_UNITS:
        return None, None
    return float(m.group(1)), unit


def parse_tab_date(ws, fallback_year, fallback_month):
    """Date from the tab's row-2 banner ('Friday, August 14, 2026').

    Falls back to the tab title ('Aug 14') plus --year/--month so a workbook
    with a reformatted banner still imports."""
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=3, values_only=True):
        for cell in row:
            m = DATE_RE.search(str(cell or ''))
            if m:
                return '%04d-%02d-%02d' % (int(m.group(3)),
                                           MONTHS.index(m.group(1).lower()) + 1,
                                           int(m.group(2)))
    m = re.search(r'([A-Za-z]{3,9})\s*(\d{1,2})', ws.title)
    if m and fallback_year:
        for i, name in enumerate(MONTHS):
            if name.startswith(m.group(1).lower()):
                return '%04d-%02d-%02d' % (fallback_year, i + 1, int(m.group(2)))
    if fallback_year and fallback_month:
        m2 = re.search(r'(\d{1,2})', ws.title)
        if m2:
            return '%04d-%02d-%02d' % (fallback_year, fallback_month, int(m2.group(1)))
    return None


def deal_identity(market_code, row):
    """Stable external-id suffix for a deal.

    Deliberately excludes the date: one deal recurring across many day tabs is
    ONE record linked to many days (that is what `day_ids` models). Includes
    the deal text so a changed price becomes a new deal rather than silently
    rewriting the old one's history."""
    basis = '|'.join([
        market_code,
        norm_key(row['brand']),
        norm_key(row['product']),
        norm_key(row['fmt']),
        norm_key(row['deal']),
    ])
    return 'ptldeal_%s_%s' % (market_code.lower(),
                              hashlib.sha1(basis.encode('utf-8')).hexdigest()[:16])


def read_rows(ws):
    """Data rows from one day tab. Section headers (Category but no Brand) and
    fully blank rows are skipped."""
    out = []
    for row in ws.iter_rows(min_row=5, max_col=7, values_only=True):
        cat, brand, product, fmt, freq, deal, loc = (list(row) + [None] * 7)[:7]
        if not (cat and brand and str(brand).strip()):
            continue
        out.append({
            'category': str(cat).strip(),
            'brand': str(brand).strip(),
            'product': str(product or '').strip(),
            'fmt': str(fmt or '').strip(),
            'freq': str(freq or '').strip(),
            'deal': re.sub(r'\s+', ' ', str(deal or '').strip()),
            'locations': str(loc or '').strip(),
        })
    return out


# ── xmlid helpers (the deal-level idempotency guard) ─────────────────────────

def load_existing_xmlids(odoo):
    recs = odoo.search_read('ir.model.data',
                            [('module', '=', IMPORT_MODULE),
                             ('model', '=', 'mint.ptl.deal')],
                            ['name', 'res_id'])
    return {r['name']: r['res_id'] for r in recs}


def bind_xmlid(odoo, name, res_id):
    odoo.create('ir.model.data', {
        'module': IMPORT_MODULE, 'name': name,
        'model': 'mint.ptl.deal', 'res_id': res_id,
    })


def resolve_day(odoo, date, market_id, cache, stats):
    """Existing (date, market) day wins — lowest id. Never creates a duplicate."""
    ck = (date, market_id)
    if ck in cache:
        return cache[ck]
    found = odoo.search('mint.ptl.day',
                        [('date', '=', date), ('market_id', '=', market_id)],
                        order='id')
    if found:
        if len(found) > 1:
            stats['duplicate_days'] += 1
            _log.warning('%s market=%s has %d day rows (%s) — using %d',
                         date, market_id, len(found), found, found[0])
        cache[ck] = found[0]
    else:
        cache[ck] = odoo.create('mint.ptl.day', {'date': date, 'market_id': market_id})
        stats['days_created'] += 1
    return cache[ck]


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--file', action='append', required=True,
                   help='PTL .xlsx (repeatable)')
    p.add_argument('--market', help='Market code (AZ/FL/IL/MI/MO/NV). '
                                    'Default: inferred from the filename.')
    p.add_argument('--year', type=int, help='Fallback year when the tab banner has no date')
    p.add_argument('--month', type=int, help='Fallback month (1-12)')
    p.add_argument('--commit', action='store_true',
                   help='Actually write. Without this nothing is created or modified.')
    p.add_argument('--limit-tabs', type=int, help='Only process the first N day tabs')
    p.add_argument('--url', default=os.environ.get('ODOO_URL', 'https://letsgomint.us'))
    p.add_argument('--db', default=os.environ.get('ODOO_DB', 'odoo'))
    p.add_argument('--uid', type=int, default=int(os.environ.get('ODOO_UID', '2')))
    p.add_argument('--key', default=os.environ.get('ODOO_KEY', ''))
    p.add_argument('-v', '--verbose', action='store_true')
    a = p.parse_args()

    logging.basicConfig(level=logging.DEBUG if a.verbose else logging.INFO,
                        format='%(levelname)-7s %(message)s')
    if not a.key:
        _log.error('No API key. Set ODOO_KEY or pass --key.')
        return 2

    odoo = Odoo(a.url, a.db, a.uid, a.key, dry_run=not a.commit)
    if not a.commit:
        _log.info('DRY-RUN — no writes. Re-run with --commit to apply.')

    regions = {r['name']: r['id'] for r in odoo.search_read('mint.region', [], ['name'])}
    _log.info('Indexing brands for exact resolution...')
    brand_idx, brand_collisions = build_brand_index(odoo)
    _log.info('  %d brand keys (%d ambiguous, dropped)',
              len(brand_idx), len(brand_collisions))

    known = load_existing_xmlids(odoo)
    _log.info('  %d deals previously imported by this script', len(known))

    stats = Counter()
    unresolved_brands, unmapped_formats = Counter(), Counter()
    day_cache, deal_cache = {}, {}
    deal_days = defaultdict(set)

    for path in a.file:
        fp = Path(path).expanduser()
        if not fp.exists():
            _log.error('missing file: %s', fp)
            return 2
        code = a.market or next((c for c in MARKET_BY_CODE if c in fp.name.upper()), None)
        if not code:
            _log.error('cannot infer market from %s — pass --market', fp.name)
            return 2
        market_id = regions.get(MARKET_BY_CODE[code])
        if not market_id:
            _log.error('mint.region %r not found', MARKET_BY_CODE[code])
            return 2
        _log.info('── %s  (market %s)', fp.name, MARKET_BY_CODE[code])

        wb = load_workbook(fp, read_only=True, data_only=True)
        tabs = wb.sheetnames[:a.limit_tabs] if a.limit_tabs else wb.sheetnames
        for tab in tabs:
            ws = wb[tab]
            date = parse_tab_date(ws, a.year, a.month)
            if not date:
                _log.warning('  tab %r: no date found — skipped', tab)
                stats['tabs_skipped'] += 1
                continue
            rows = read_rows(ws)
            if not rows:
                continue
            stats['tabs'] += 1
            day_id = resolve_day(odoo, date, market_id, day_cache, stats)

            for row in rows:
                stats['rows'] += 1
                xmlid = deal_identity(code, row)

                bid = brand_idx.get(norm_key(row['brand']))
                if bid:
                    stats['brand_resolved'] += 1
                else:
                    unresolved_brands[row['brand']] += 1

                wval, wunit = parse_weight(row['product'])
                if wval:
                    stats['weight_parsed'] += 1

                if row['product']:
                    stats['product_descriptor'] += 1

                dtype = FORMAT_TO_DISCOUNT_TYPE.get(norm_key(row['fmt']))
                if row['fmt'] and not dtype:
                    unmapped_formats[row['fmt']] += 1

                name = ' '.join(x for x in (row['deal'], row['brand'], row['product']) if x)[:200]
                vals = {
                    'name': name or row['brand'],
                    'market_id': market_id,
                    'display_text': row['deal'][:200] or None,
                    # Product-form descriptor ("Flower 3.5g"); see NOTE ON PRODUCTS.
                    'description': row['product'] or None,
                    'product_category': row['category'] if row['category'] in (
                        'Flower', 'Pre-Rolls', 'Vapes', 'Edibles & Tinctures',
                        'Concentrates & Topicals', 'Featured Deals') else None,
                }
                if bid:
                    vals['brand_id'] = bid
                if wval:
                    vals['weight_value'], vals['weight_unit'] = wval, wunit
                if dtype:
                    vals['discount_type'] = dtype

                if xmlid in deal_cache:
                    deal_id = deal_cache[xmlid]
                elif xmlid in known:
                    deal_id = known[xmlid]
                    odoo.write('mint.ptl.deal', [deal_id], vals)
                    deal_cache[xmlid] = deal_id
                    stats['deals_updated'] += 1
                else:
                    vals['state'] = 'pending'      # never auto-publish to Dutchie
                    deal_id = odoo.create('mint.ptl.deal', vals)
                    bind_xmlid(odoo, xmlid, deal_id)
                    deal_cache[xmlid] = deal_id
                    stats['deals_created'] += 1
                deal_days[deal_id].add(day_id)
        wb.close()

    # Link deals to every day tab they appeared on (one write per deal).
    for deal_id, days in deal_days.items():
        odoo.write('mint.ptl.deal', [deal_id], {'day_ids': [(4, d) for d in sorted(days)]})
        stats['day_links'] += len(days)

    print('\n' + '=' * 66)
    print('PTL spreadsheet import — %s' % ('COMMITTED' if a.commit else 'DRY RUN'))
    print('=' * 66)
    for k in ('tabs', 'tabs_skipped', 'rows', 'days_created', 'duplicate_days',
              'deals_created', 'deals_updated', 'day_links',
              'brand_resolved', 'weight_parsed', 'product_descriptor'):
        print(f'  {k:<20} {stats[k]}')
    if stats['rows']:
        print(f'\n  brand   resolved: {stats["brand_resolved"]/stats["rows"]*100:5.1f}%')
        print(f'  weight  parsed  : {stats["weight_parsed"]/stats["rows"]*100:5.1f}%')
        print(f'  product descr.  : {stats["product_descriptor"]/stats["rows"]*100:5.1f}%')
    if unresolved_brands:
        print(f'\n  UNRESOLVED BRANDS ({len(unresolved_brands)} distinct) — '
              f'add these as mint.brand.aliases, do not guess:')
        for b, n in unresolved_brands.most_common(25):
            print(f'     {n:>5}x  {b}')
    if unmapped_formats:
        print(f'\n  UNMAPPED FORMATS ({len(unmapped_formats)} distinct) — '
              f'left without a discount_type:')
        for f, n in unmapped_formats.most_common(15):
            print(f'     {n:>5}x  {f}')
    print()
    return 0


if __name__ == '__main__':
    sys.exit(main())
