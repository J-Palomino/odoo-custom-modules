#!/usr/bin/env python3
"""
backfill_brands_outreach_status.py — populate mint.national.promo.contact_status,
ready_to_plot, and plotted from the Brands.xlsx workbook.

Reads:
    ~/Downloads/National Promos by brand/Brands.xlsx

Sheet shapes vary; the script handles the common ones:
  - "Brands Reached out to 02.20.2026"-style outreach tracker
  - "AZ", "MO", "NV", "IL" — per-state plotted-status (Inventory Brand Name | Ready | Plotted)
  - "AZ April", "AZ May", etc. — per-state, per-month plotted-status

For each row that has a brand name and a state, find the matching
`mint.national.promo(brand, market, year=2026)` and set:
  - contact_status: from the outreach status column (Got Confirmation / Waiting for Answer)
  - ready_to_plot:  True if "YES" in any Ready column
  - plotted:        True if "YES" in any Plotted column

Idempotent.
"""
import argparse
import json
import logging
import os
import sys
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
_log = logging.getLogger(__name__)


def norm_brand(s: str) -> str:
    return ''.join(c for c in (s or '').lower() if c.isalnum())


def status_from_text(text: str):
    """Map outreach text → mint.national.promo.contact_status selection."""
    if not text:
        return None
    t = str(text).strip().lower()
    if 'got confirmation' in t or 'confirmed' in t or 'yes' == t:
        return 'confirmed'
    if 'waiting' in t or 'await' in t:
        return 'awaiting_response'
    if 'declined' in t or 'no thanks' in t:
        return 'declined'
    if 'outreach' in t or 'reached out' in t:
        return 'outreached'
    return None


class Odoo:
    def __init__(self, url, db, uid, key, dry_run=False):
        self.url = url.rstrip('/')
        if not self.url.endswith('/jsonrpc'):
            self.url = self.url + '/jsonrpc'
        self.db = db
        self.uid = uid
        self.key = key
        self.dry_run = dry_run
        self.headers = {
            'Content-Type': 'application/json',
            'User-Agent': 'mint-backfill/1.0',
        }

    def call(self, model, method, args, kwargs=None):
        if self.dry_run and method in ('create', 'write'):
            _log.info('[DRY-RUN] %s.%s args=%s kw=%s', model, method, args, kwargs)
            return {'result': True}
        payload = {
            'jsonrpc': '2.0',
            'method': 'call',
            'params': {
                'service': 'object',
                'method': 'execute_kw',
                'args': [self.db, self.uid, self.key, model, method, args, kwargs or {}],
            },
        }
        req = urllib.request.Request(
            self.url,
            data=json.dumps(payload).encode('utf-8'),
            headers=self.headers,
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read())
        if 'error' in data:
            raise RuntimeError(data['error'].get('data', {}).get('message', data['error']))
        return data

    def search_read(self, model, domain, fields, **kw):
        return self.call(model, 'search_read', [domain], {'fields': fields, **kw}).get('result', [])

    def write(self, model, ids, vals):
        return self.call(model, 'write', [ids, vals]).get('result')


def load_brand_lookup(odoo):
    out = {}
    offset = 0
    while True:
        page = odoo.search_read('mint.brand', [], ['id', 'name'], limit=500, offset=offset)
        if not page:
            break
        for rec in page:
            n = norm_brand(rec['name'])
            if n and n not in out:
                out[n] = rec['id']
        offset += 500
    return out


def load_market_lookup(odoo):
    regs = odoo.search_read('mint.region', [], ['id', 'code'])
    return {r['code'].upper(): r['id'] for r in regs if r.get('code')}


def find_campaign(odoo, brand_id, market_id):
    rows = odoo.search_read(
        'mint.national.promo',
        [['brand_id', '=', brand_id], ['market_id', '=', market_id], ['year', '=', 2026]],
        ['id'],
        limit=1,
    )
    return rows[0]['id'] if rows else None


def process_state_sheet(ws, state_code, brand_lookup, market_id, odoo):
    """Read per-state outreach sheet rows. Returns (updated, unmatched, no_campaign)."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return 0, 0, 0

    # Heuristic: brand column is col 1 ("Inventory Brand Name"); ready/plotted in cols 2-5
    updated = 0
    unmatched = 0
    no_campaign = 0
    for row in rows[2:]:
        if not row or len(row) < 2:
            continue
        brand_text = row[1] if isinstance(row[1], str) else None
        if not brand_text or brand_text.lower().startswith('inventory'):
            continue
        bid = brand_lookup.get(norm_brand(brand_text))
        if not bid:
            unmatched += 1
            sys.stderr.write(f'unmatched\t{state_code}\t{brand_text}\n')
            continue
        campaign_id = find_campaign(odoo, bid, market_id)
        if not campaign_id:
            no_campaign += 1
            continue
        ready = any(
            isinstance(v, str) and v.strip().upper() == 'YES' for v in row[2:6]
        )
        # plotted columns typically appear to the right of ready columns
        plotted = any(
            isinstance(v, str) and v.strip().upper() == 'YES' for v in row[4:10]
        )
        vals = {}
        if ready:
            vals['ready_to_plot'] = True
        if plotted:
            vals['plotted'] = True
        if vals:
            odoo.write('mint.national.promo', [campaign_id], vals)
            updated += 1
    return updated, unmatched, no_campaign


def process_outreach_sheet(ws, brand_lookup, market_lookup, odoo):
    """Read the 'Brands Reached out to ...' tracker. Returns updates count.

    Layout: column blocks (state header in row 4), brand name in col below, optional
    status note in adjacent col. Multiple state blocks across columns.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    # Find header row containing state codes
    header_row_idx = None
    for i, row in enumerate(rows[:10]):
        cells = [str(v).strip().upper() if isinstance(v, str) else None for v in row]
        if any(c in ('AZ', 'MO', 'NV', 'IL') for c in cells if c):
            header_row_idx = i
            break
    if header_row_idx is None:
        return 0

    state_cols = {}  # state_code → col_idx
    for col_idx, val in enumerate(rows[header_row_idx]):
        if isinstance(val, str) and val.strip().upper() in ('AZ', 'MO', 'NV', 'IL'):
            state_cols[val.strip().upper()] = col_idx

    updated = 0
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        for state_code, col_idx in state_cols.items():
            if col_idx >= len(row):
                continue
            brand_text = row[col_idx]
            if not isinstance(brand_text, str) or not brand_text.strip():
                continue
            note = row[col_idx + 1] if col_idx + 1 < len(row) else None
            status = status_from_text(note) or status_from_text(brand_text)

            bid = brand_lookup.get(norm_brand(brand_text))
            if not bid:
                continue
            market_id = market_lookup.get(state_code)
            if not market_id:
                continue
            campaign_id = find_campaign(odoo, bid, market_id)
            if not campaign_id:
                continue
            vals = {'contact_status': status} if status else {}
            if status == 'confirmed':
                vals['ready_to_plot'] = True
            if vals:
                odoo.write('mint.national.promo', [campaign_id], vals)
                updated += 1
    return updated


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--xlsx', default=str(
        Path.home() / 'Downloads' / 'National Promos by brand' / 'Brands.xlsx',
    ))
    p.add_argument('--dry-run', action='store_true')
    p.add_argument('--url', default=os.environ.get('ODOO_URL', 'https://letsgomint.us/jsonrpc'))
    p.add_argument('--db', default=os.environ.get('ODOO_DB', 'odoo'))
    p.add_argument('--uid', type=int, default=int(os.environ.get('ODOO_UID', '2')))
    p.add_argument('--key', default=os.environ.get('ODOO_KEY', ''))
    args = p.parse_args()

    if not args.key:
        sys.exit('ODOO_KEY env var or --key required')

    odoo = Odoo(args.url, args.db, args.uid, args.key, dry_run=args.dry_run)
    _log.info('Loading mint.brand registry...')
    brand_lookup = load_brand_lookup(odoo)
    _log.info('  %d brands', len(brand_lookup))
    market_lookup = load_market_lookup(odoo)

    wb = load_workbook(args.xlsx, data_only=True)
    sys.stderr.write('unmatched_brand_log:\tcontext\tbrand_text\n')

    total_updated = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        upper = sheet_name.strip().upper()
        # Per-state sheets: "AZ", "MO", "NV", "IL", "AZ April", "AZ May", etc.
        first_token = upper.split()[0] if upper else ''
        if first_token in ('AZ', 'MO', 'NV', 'IL') and first_token in market_lookup:
            updated, unmatched, no_campaign = process_state_sheet(
                ws, first_token, brand_lookup, market_lookup[first_token], odoo,
            )
            _log.info('sheet %r: updated=%d unmatched=%d no_campaign=%d',
                      sheet_name, updated, unmatched, no_campaign)
            total_updated += updated
        elif 'REACHED OUT' in upper:
            updated = process_outreach_sheet(ws, brand_lookup, market_lookup, odoo)
            _log.info('sheet %r (outreach): updated=%d', sheet_name, updated)
            total_updated += updated

    _log.info('done. total_updated=%d', total_updated)


if __name__ == '__main__':
    main()
